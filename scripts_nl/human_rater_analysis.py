from typing import List, Dict
from collections import Counter, defaultdict
from openpyxl import Workbook
import json
import statistics


def complete_columns(sentence: Dict, headers: List[str]) -> None:
    for annotation in sentence["annotations"]:
        if "num_modifications" not in annotation:
            annotation["num_modifications"] = 0
            annotation["modified"] = None
        else:
            annotation["num_modifications"] = int(annotation["num_modifications"])
        if annotation["unanswerable"]:
            complete_unanswerable(annotation, headers)
        for header in annotation:
            if "_scale" in header and annotation[header] != "NA":
                annotation[header] = int(annotation[header])


def complete_unanswerable(annotation: Dict, headers: List[str]) -> None:
    for header in headers:
        if header not in annotation:
            annotation[header] = "NA"


def calculate_avg_rater_score(impact_scale: str, sentence: dict, avg_type: str = "mean") -> float:
    scores = [anno[impact_scale] for anno in sentence["annotations"] if isinstance(anno[impact_scale], int)]
    if avg_type == "mean":
        return statistics.mean(scores)
    elif avg_type == "median":
        return statistics.median(scores)
    elif avg_type == "mode":
        return statistics.mode(scores)
    else:
        raise ValueError("avg_type must be one of 'mean', 'median' or 'mode'")


# We should report agreement using multiple theoretical null-distributions:
# - uniform null: each point on the scale is equally likely to be chosen -> expected variance = (A^2 - 1) / 12 = 2
# - normal: raters avoid extremes, so there is a central tendency -> expected variance = 1.04
# - maximum dissensus: raters always choose extremes -> expected variance = 4
# We use a five point Likert scale
def calculate_sentence_interrater_agreement(rater_scores: List[int], null_dist: str) -> float:
    if null_dist == 'uniform':
        expected_var = 2
    elif null_dist == 'normal':     # See LeBreton and Senter (2008)
        expected_var = 1.04
    elif null_dist == 'triangular':
        # See LeBreton and Senter (2008)
        # rating       1   2   3   4   5
        # probability .11 .22 .34 .22 .11
        # rating 3 is twice as likely as ratings 2 and 4, and thrice as likely as ratings 1 and 5
        expected_var = 1.32
    elif null_dist == 'inverse_triangular':
        # Inverse triangular based on triangular distribution in LeBreton and Senter
        # rating 3 is half as likely as ratings 2 and 4, and one third as likely as ratings 1 and 5
        # One rating of 3 corresponds to 2 ratings of 2, 2 of 4, and three of 1 and 5 each
        # Totals 11 ratings, so the following probabilities
        # rating 1 = 3 / 11 = 0.2727
        # rating 2 = 2 / 11 = 0.1818
        # rating 3 = 1 / 11 = 0.0909
        # rating 4 = 2 / 11 = 0.1818
        # rating 5 = 3 / 11 = 0.2727
        # rating        1     2     3     4     5
        # Mean     =  1 * 3 / 11 + 2 * 2 / 11 + 3 * 1 / 11 + 4 * 2 / 11 + 5 * 3 / 11
        #          = 33 / 11
        #          = 3
        # Variance = 2^2 * 3/11 + 1^2 * 2/11 + 0^2 * 3/11 + 1^2 * 2/11 + 2^2 * 3/11
        # Variance = 12/11 + 2/11 + 0/11 + 2/11 + 12/11
        #          = 28/11
        #          = 2.5454
        expected_var = 28/11
    elif null_dist == 'max_dissensus':
        # σ_{mv}^2 = 0.5(X_U^2 + X_L^2)−[0.5(X_U + X_L)]2
        #          = 0.5(5^2 + 1^2) - [0.5(5+1)]^2
        #          = 0.5(26) - [0.5(6)]^2
        #          = 13 - 3^2
        #          = 4
        expected_var = 4
    else:
        raise ValueError('"null_dist" must be "normal", "uniform" or "max_dissensus"')
    # The raters are the population, not a sample, so use population variance
    var = statistics.pvariance(rater_scores)
    # r^∗_{wg} = 1 − ( S_x^2 / σ^2 )
    ira = 1 - (var / expected_var)
    return ira


def get_sentences_high_ira(sentences: List[dict], impact_scale: str, ira_threshold: float, null_dist: str) -> List[dict]:
    filtered = [sentence for sentence in sentences if len(get_rater_scores(sentence, impact_scale)) > 1]
    return [sentence for sentence in filtered if get_sentence_ira(sentence, impact_scale, null_dist) >= ira_threshold]


def get_sentences_low_ira(sentences: List[dict], impact_scale: str, ira_threshold: float, null_dist: str) -> List[dict]:
    filtered = [sentence for sentence in sentences if len(get_rater_scores(sentence, impact_scale)) > 1]
    return [sentence for sentence in filtered if get_sentence_ira(sentence, impact_scale, null_dist) < ira_threshold]


def get_sentence_ira(sentence: dict, impact_scale: str, null_dist: str) -> float:
    scores = get_rater_scores(sentence, impact_scale)
    return calculate_sentence_interrater_agreement(scores, null_dist)


def get_ira_scores(sentence_ratings: List[dict], impact_scale: str, null_dist) -> List[float]:
    return [get_sentence_ira(sentence, impact_scale, null_dist) for sentence in sentence_ratings]


def get_ira_dist(sentence_ratings: List[dict], null_dist: str):
    impact_scales = ["emotional_scale", "style_scale", "reflection_scale", "narrative_scale"]
    ira_dist = defaultdict(Counter)
    for impact_scale in impact_scales:
        sentences_rated = [sent for sent in sentence_ratings if len(get_rater_scores(sent, impact_scale)) > 1]
        ira_scores = get_ira_scores(sentences_rated, impact_scale, null_dist)
        print(f'\t{impact_scale: <20}\tmean IRA:', statistics.mean(ira_scores), '\tstdev IRA:', statistics.stdev(ira_scores))
        for ira_score in ira_scores:
            ira_dist[impact_scale].update([get_ira_range(ira_score)])
        #for ira_range, freq in sorted(ira_dist[impact_scale].items(), key=lambda x: x[0]):
        #    print(impact_scale, ira_range, freq)
    return ira_dist


# From O'Neil (2017): For application to the rwg family, the following standards were recommended:
# 0–0.30 (lack of agreement),
# 0.31–0.50 (weak agreement),
# 0.51–0.70 (moderate agreement),
# 0.71–0.90 (strong agreement),
# 0.91–1.0 (very strong agreement).
def get_ira_range(ira_score: float) -> str:
    score = abs(ira_score)
    if ira_score < -0.5:
        return '-0.70 -- -0.51'
    elif ira_score < -0.3:
        return '-0.50 -- -0.31'
    elif ira_score < 0.0:
        return '-0.30 -- 0.00'
    elif ira_score <= 0.3:
        return '0.00 -- +0.30'
    elif ira_score <= 0.5:
        return '+0.31 -- +0.50'
    elif ira_score <= 0.7:
        return '+0.51 -- +0.70'
    elif ira_score <= 0.9:
        return '+0.71 -- +0.90'
    else:
        return '+0.91 -- +1.00'


def get_rater_scores(sentence: dict, impact_scale: str) -> List[int]:
    return [anno[impact_scale] for anno in sentence["annotations"] if isinstance(anno[impact_scale], int)]


def parse_annotations(sentence: dict, headers: List[str], sheet, annotation_row_num: int) -> None:
    for annotation in sentence["annotations"]:
        set_base_cells(sentence, sheet, annotation_row_num)
        sheet.cell(column=3, row=annotation_row_num, value=annotation["annotator"])
        column_num = 4
        for header in headers[3:]:
            sheet.cell(column=column_num, row=annotation_row_num, value=annotation[header])
            column_num += 1
        annotation_row_num += 1


def set_base_cells(sentence: dict, sheet, row_num: int) -> None:
    sheet.cell(column=1, row=row_num, value=sentence["sentence_id"])
    sheet.cell(column=2, row=row_num, value=sentence["text"])


def parse_impact_scale(impact_scale: str, sentence: dict, sheet, row_num: int) -> None:
    set_base_cells(sentence, sheet, row_num)
    num_annotators = len(sentence["annotations"])
    total_score = sum([anno[impact_scale] for anno in sentence["annotations"] if isinstance(anno[impact_scale], int)])
    avg_score = total_score / num_annotators
    has_na = len([anno[impact_scale] for anno in sentence["annotations"] if anno[impact_scale] == "NA"]) > 0
    for anno_index, anno in enumerate(sentence["annotations"]):
        col = anno_index + 3
        sheet.cell(column=col, row=row_num, value=anno[impact_scale])
    rule_based_score = 0
    if impact_scale in sentence["model_impact_score"]:
        rule_based_score = sentence["model_impact_score"][impact_scale]
    sheet.cell(column=8, row=row_num, value=num_annotators)
    sheet.cell(column=9, row=row_num, value=avg_score)
    sheet.cell(column=10, row=row_num, value=has_na)
    sheet.cell(column=11, row=row_num, value=rule_based_score)


def clear_unanswerable_scales(sentence: dict) -> None:
    impact_scales = ["emotional_scale", "style_scale", "reflection_scale", "narrative_scale", "emotional_valence"]
    for annotation in sentence['annotations']:
        if annotation['unanswerable'] is False:
            continue
        for impact_scale in impact_scales:
            if impact_scale in annotation:
                del annotation[impact_scale]


def get_sentence_ratings(data_file: str) -> List[dict]:
    with open(data_file, 'rt') as fh:
        data = json.load(fh)
    sentence_ratings = [sentence_doc["_source"] for sentence_doc in data]
    for sentence in sentence_ratings:
        clear_unanswerable_scales(sentence)
    return sentence_ratings


def write_rating_spreadsheet(sentence_ratings: List[dict], config: dict) -> None:
    wb = Workbook()
    sheets = {
        "annotations": wb.active,
    }

    impact_scales = ["emotional_scale", "style_scale", "reflection_scale", "narrative_scale", "emotional_valence"]
    for impact_scale in impact_scales:
        sheets[impact_scale] = wb.create_sheet(title=impact_scale)
        sheets[impact_scale].append(
            ["sentence_id", "text", "annotator1", "annotator2", "annotator3", "annotator4", "annotator5",
             "num_annotators", "avg_score", "has_NA", "impact_model_score"])

    headers = [
        'sentence_id',
        'text',
        'annotator',
        'emotional_scale',
        'emotional_valence',
        'style_scale',
        'reflection_scale',
        'narrative_scale',
        'unanswerable',
        'num_modifications',
        'created',
        'modified',
    ]

    wb.active.append(headers)

    header_row = ["sentence_id"]
    for header in headers:
        if header == "sentence_id":
            continue
        for i in range(1, 4):
            header_row += ["{}_{}".format(header, i)]

    impact_row_num = 2
    annotation_row_num = 2
    for sentence in sentence_ratings:
        complete_columns(sentence, headers)
        for impact_scale in impact_scales:
            sheet = sheets[impact_scale]
            parse_impact_scale(impact_scale, sentence, sheet, impact_row_num)
        parse_annotations(sentence, headers, sheets["annotations"], annotation_row_num)
        annotation_row_num += len(sentence["annotations"])
        impact_row_num += 1

    print(f'\twriting ratings to spreadsheet file {config["spreadsheet_file"]}')
    wb.save(filename=config['spreadsheet_file'])
