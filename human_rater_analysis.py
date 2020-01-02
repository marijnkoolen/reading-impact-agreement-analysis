from typing import List, Dict, Union
from openpyxl import Workbook
import json
import statistics


def complete_columns(sentence: Dict, headers: List[str]) -> None:
    for annotation in sentence["annotations"]:
        if "num_modifications" not in annotation:
            annotation["num_modifications"] = 0
            annotation["modified"] = None
        else:
            annotation["num_modifications"] = int(annotation["num_modifications"])
        if annotation["unanswerable"]:
            complete_unanswerable(annotation, headers)
        for header in annotation:
            if "_scale" in header and annotation[header] != "NA":
                annotation[header] = int(annotation[header])


def complete_unanswerable(annotation: Dict, headers: List[str]) -> None:
    for header in headers:
        if header not in annotation:
            annotation[header] = "NA"


def calculate_avg_rater_score(impact_scale: str, sentence: dict, avg_type: str = "mean") -> float:
    scores = [anno[impact_scale] for anno in sentence["annotations"] if isinstance(anno[impact_scale], int)]
    if avg_type == "mean":
        return statistics.mean(scores)
    elif avg_type == "median":
        return statistics.median(scores)
    elif avg_type == "mode":
        return statistics.mode(scores)
    else:
        raise ValueError("avg_type must be one of 'mean', 'median' or 'mode'")


def calculcate_inter_rater_agreement(rater_scores: List[int]) -> float:
    # σ^2 * mv = 0.5(X_U^2 + X_L^2)−[0.5(X_U + X_L)]2
    #          = 0.5(5^2 + 1^2) - [0.5(5+1)]^2
    #          = 0.5(26) - [0.5(6)]^2
    #          = 13 - 3^2
    #          = 4
    max_dissensus = 4
    var = statistics.variance(rater_scores)
    # r^∗_{wg}(j)=1−(S_x^2/σ_{mv}^2)
    ira_score = 1 - (var / max_dissensus)
    return ira_score


def get_sentences_high_ira(sentences: List[dict], impact_scale: str, ira_threshold: float) -> List[dict]:
    filtered = [sentence for sentence in sentences if len(get_rater_scores(sentence, impact_scale)) > 1]
    return [sentence for sentence in filtered if get_sentence_ira(sentence, impact_scale) >= ira_threshold]


def get_sentences_low_ira(sentences: List[dict], impact_scale: str, ira_threshold: float) -> list[dict]:
    filtered = [sentence for sentence in sentences if len(get_rater_scores(sentence, impact_scale)) > 1]
    return [sentence for sentence in filtered if get_sentence_ira(sentence, impact_scale) < ira_threshold]


def get_sentence_ira(sentence: dict, impact_scale: str) -> float:
    scores = get_rater_scores(sentence, impact_scale)
    return calculcate_inter_rater_agreement(scores)


def get_rater_scores(sentence: dict, impact_scale: str) -> List[int]:
    return [anno[impact_scale] for anno in sentence["annotations"] if isinstance(anno[impact_scale], int)]


def parse_annotations(sentence: dict, headers: List[str], sheet, annotation_row_num: int) -> None:
    for annotation in sentence["annotations"]:
        set_base_cells(sentence, sheet, annotation_row_num)
        sheet.cell(column=3, row=annotation_row_num, value=annotation["annotator"])
        column_num = 4
        for header in headers[3:]:
            sheet.cell(column=column_num, row=annotation_row_num, value=annotation[header])
            column_num += 1
        annotation_row_num += 1


def set_base_cells(sentence: dict, sheet, row_num: int) -> None:
    sheet.cell(column=1, row=row_num, value=sentence["sentence_id"])
    sheet.cell(column=2, row=row_num, value=sentence["text"])


def parse_impact_scale(impact_scale: str, sentence: dict, sheet, row_num: int) -> None:
    set_base_cells(sentence, sheet, row_num)
    num_annotators = len(sentence["annotations"])
    total_score = sum([anno[impact_scale] for anno in sentence["annotations"] if isinstance(anno[impact_scale], int)])
    avg_score = total_score / num_annotators
    has_na = len([anno[impact_scale] for anno in sentence["annotations"] if anno[impact_scale] == "NA"]) > 0
    for anno_index, anno in enumerate(sentence["annotations"]):
        col = anno_index + 3
        sheet.cell(column=col, row=row_num, value=anno[impact_scale])
    rule_based_score = 0
    if impact_scale in sentence["model_impact_score"]:
        rule_based_score = sentence["model_impact_score"][impact_scale]
    sheet.cell(column=8, row=row_num, value=num_annotators)
    sheet.cell(column=9, row=row_num, value=avg_score)
    sheet.cell(column=10, row=row_num, value=has_na)
    sheet.cell(column=11, row=row_num, value=rule_based_score)


def get_sentence_ratings(data_file: str) -> List[dict]:
    with open(data_file, 'rt') as fh:
        data = json.load(fh)
    return [sentence_doc["_source"] for sentence_doc in data]


def write_rating_spreadsheet(sentence_ratings: List[dict]) -> None:
    wb = Workbook()
    sheets = {
        "annotations": wb.active,
    }

    impact_scales = ["emotional_scale", "style_scale", "reflection_scale", "narrative_scale", "emotional_valence"]
    for impact_scale in impact_scales:
        sheets[impact_scale] = wb.create_sheet(title=impact_scale)
        sheets[impact_scale].append(
            ["sentence_id", "text", "annotator1", "annotator2", "annotator3", "annotator4", "annotator5",
             "num_annotators", "avg_score", "has_NA", "impact_model_score"])

    headers = [
        'sentence_id',
        'text',
        'annotator',
        'emotional_scale',
        'emotional_valence',
        'style_scale',
        'reflection_scale',
        'narrative_scale',
        'unanswerable',
        'num_modifications',
        'created',
        'modified',
    ]

    wb.active.append(headers)

    header_row = ["sentence_id"]
    for header in headers:
        if header == "sentence_id":
            continue
        for i in range(1, 4):
            header_row += ["{}_{}".format(header, i)]

    impact_row_num = 2
    annotation_row_num = 2
    for sentence in sentence_ratings:
        complete_columns(sentence, headers)
        for impact_scale in impact_scales:
            sheet = sheets[impact_scale]
            parse_impact_scale(impact_scale, sentence, sheet, impact_row_num)
        parse_annotations(sentence, headers, sheets["annotations"], annotation_row_num)
        annotation_row_num += len(sentence["annotations"])
        impact_row_num += 1

    output_file = "reading_impact_questionnaire_data.xlsx"
    wb.save(filename=output_file)
